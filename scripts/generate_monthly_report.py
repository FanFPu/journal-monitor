#!/usr/bin/env python3

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import re
import sys
import time
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font


USER_AGENT = "CancerMultiomicsJournalMonitor/1.0 (+Codex)"
SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
REGISTRY_PATH = SKILL_DIR / "references" / "journal_registry.json"
DEFAULT_OUTPUT_DIR = Path.cwd() / "reports"

CANCER_PATTERNS = {
    "乳腺癌": [r"\bbreast cancer\b", r"\blobular breast\b"],
    "肺癌": [r"\blung cancer\b", r"\blung adenocarcinoma\b", r"\bnsclc\b", r"\bsclc\b"],
    "胃癌": [r"\bgastric cancer\b", r"\bstomach cancer\b"],
    "卵巢癌": [r"\bovarian cancer\b", r"\bovarian carcinoma\b", r"\bhigh-grade serous\b"],
    "黑色素瘤": [r"\bmelanoma\b"],
    "胶质瘤/脑肿瘤": [r"\bglioblastoma\b", r"\bglioma\b", r"\bmedulloblastoma\b", r"\bbrain metasta"],
    "间皮瘤": [r"\bmesothelioma\b"],
    "白血病": [r"\bleukemia\b", r"\baml\b", r"\bacute myeloid leukemia\b"],
    "结直肠癌": [r"\bcolorectal cancer\b", r"\bcolon cancer\b", r"\brectal cancer\b"],
    "胆道肿瘤": [r"\bcholangiocarcinoma\b", r"\bbiliary tract cancer\b"],
    "胰腺癌": [r"\bpancreatic cancer\b", r"\bpdac\b"],
    "鼻咽癌": [r"\bnasopharyngeal carcinoma\b"],
    "前列腺癌": [r"\bprostate cancer\b"],
    "宫颈癌": [r"\bcervical cancer\b"],
    "肉瘤": [r"\bsarcoma\b", r"\bdesmoplastic small round cell tumor\b"],
    "泛癌": [r"\bpancancer\b", r"\bpan-cancer\b", r"\bmultiple solid tumors\b", r"\bsolid tumors\b"],
}

STRONG_CANCER_PATTERNS = [
    r"\bcancer\b",
    r"\btumou?r\b",
    r"\bcarcinoma\b",
    r"\bneoplasm\b",
    r"\bsarcoma\b",
    r"\bleukemia\b",
    r"\blymphoma\b",
    r"\bmelanoma\b",
    r"\bglioma\b",
    r"\bmedulloblastoma\b",
    r"\bmyeloma\b",
    r"\bmesothelioma\b",
    r"\bmetastasis\b",
    r"\badenocarcinoma\b",
    r"\bmalignan",
]

OMICS_PATTERNS = {
    "genomics": [r"\bgenomic", r"\bgenome", r"\bwhole[- ]genome", r"\bwhole[- ]exome", r"\bwgs\b", r"\bwes\b", r"\bcopy number", r"\bcnv\b"],
    "epigenomics": [r"\bepigen", r"\bdna methyl", r"\bmethylome", r"\bchromatin", r"\batac[- ]seq", r"\bchip[- ]seq"],
    "transcriptomics": [r"\btranscriptom", r"\brna[- ]seq", r"\bscrna", r"\bsnrna", r"\bsingle[- ]cell", r"\bspatial transcript", r"\bgene expression"],
    "proteomics": [r"\bproteom", r"\bphosphoproteom", r"\bmass spectrometry", r"\bimmunopeptidome"],
    "metabolomics": [r"\bmetabolom", r"\blipidom", r"\blipidomic", r"\bmetabolic profiling"],
    "microbiome": [r"\bmicrobiome", r"\bmicrobiota", r"\bmetagenom"],
    "multiomics": [r"\bmulti[- ]omics\b", r"\bmultiomics\b", r"\bmulti[- ]omic\b"],
}

EXPLICIT_MULTIOMICS_PATTERNS = [
    r"\bmulti[- ]omics\b",
    r"\bmultiomics\b",
    r"\bmulti[- ]omic\b",
    r"\bproteogenomic",
    r"\bspatial multi[- ]omics\b",
    r"\bsingle[- ]cell multiomics\b",
    r"\bintegrated multiomic",
    r"\bintegrated multi-omic",
    r"\bmulti-layered molecular profiling\b",
]

REVIEW_PATTERNS = [
    r"\breview\b",
    r"\bperspective\b",
    r"\bcommentary\b",
    r"\bnews\b",
    r"\beditorial\b",
]

ACCESSION_PATTERNS = [
    r"\bGSE\d+\b",
    r"\bGSM\d+\b",
    r"\bGPL\d+\b",
    r"\bGDS\d+\b",
    r"\bSRP\d+\b",
    r"\bSRX\d+\b",
    r"\bSRR\d+\b",
    r"\bSRS\d+\b",
    r"\bERP\d+\b",
    r"\bERX\d+\b",
    r"\bERR\d+\b",
    r"\bPRJNA\d+\b",
    r"\bPRJEB\d+\b",
    r"\bEGAS\d+\b",
    r"\bEGAD\d+\b",
    r"\bphs\d+(?:\.v\d+)?\b",
    r"\bPXD\d+\b",
    r"\bE-MTAB-\d+\b",
]

SAMPLE_PATTERNS = [
    r"\b\d{1,4}\s+patients?\b",
    r"\b\d{1,4}\s+patient[s]?\b",
    r"\b\d{1,4}\s+cases?\b",
    r"\b\d{1,4}\s+tumou?rs?\b",
    r"\b\d{1,4}\s+samples?\b",
    r"\b\d{1,4}\s+lesions?\b",
    r"\b\d{1,4}\s+cohorts?\b",
    r"\bcohort of \d{1,4}\b",
    r"\b\d{1,4}\s+participants?\b",
    r"\b\d{1,4}(?:,\d{3})*\s+cells?\b",
]


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def load_registry() -> List[Dict[str, str]]:
    return json.loads(REGISTRY_PATH.read_text())


def today_utc() -> dt.date:
    return dt.datetime.now().date()


def build_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    return session


def chunked(items: List[str], size: int) -> List[List[str]]:
    return [items[index:index + size] for index in range(0, len(items), size)]


def pubmed_search(session: requests.Session, journal_name: str, start_date: dt.date, end_date: dt.date) -> List[str]:
    multiomics_clause = " OR ".join(
        [
            '"multi-omics"',
            "multiomics",
            '"multi-omic"',
            '"single-cell multiomics"',
            '"spatial multi-omics"',
            "proteogenomic",
            "proteogenomics",
            '"molecular profiling"',
            '"multi-layered molecular profiling"',
            '"integrated spatial analysis"',
            "transcriptomic",
            "proteomic",
            "metabolomic",
            "lipidomic",
            '"DNA methylation"',
            '"spatial transcriptomics"',
        ]
    )
    cancer_clause = " OR ".join(
        [
            "cancer",
            "tumor",
            "tumour",
            "carcinoma",
            "neoplasm",
            "sarcoma",
            "leukemia",
            "lymphoma",
            "melanoma",
            "glioma",
            "metastasis",
            "oncology",
        ]
    )
    term = (
        f'"{journal_name}"[jour] AND '
        f'("{start_date:%Y/%m/%d}"[Date - Publication] : "{end_date:%Y/%m/%d}"[Date - Publication]) AND '
        f'(({multiomics_clause}) AND ({cancer_clause}))'
    )
    response = session.get(
        "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi",
        params={"db": "pubmed", "retmode": "json", "retmax": 200, "term": term},
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()
    return payload.get("esearchresult", {}).get("idlist", [])


def pubmed_fetch_records(session: requests.Session, pmids: List[str]) -> List[Dict[str, object]]:
    records: List[Dict[str, object]] = []
    for batch in chunked(pmids, 50):
        response = session.get(
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi",
            params={"db": "pubmed", "retmode": "xml", "id": ",".join(batch)},
            timeout=60,
        )
        response.raise_for_status()
        root = ET.fromstring(response.content)
        for article in root.findall(".//PubmedArticle"):
            records.append(parse_pubmed_article(article))
        time.sleep(0.2)
    return records


def parse_pubmed_article(article: ET.Element) -> Dict[str, object]:
    pmid = article.findtext(".//PMID", default="")
    title_node = article.find(".//ArticleTitle")
    title = normalize_space("".join(title_node.itertext()) if title_node is not None else "")
    abstract_parts = []
    for node in article.findall(".//Abstract/AbstractText"):
        text = normalize_space("".join(node.itertext()))
        if text:
            label = node.attrib.get("Label")
            abstract_parts.append(f"{label}: {text}" if label else text)
    abstract = " ".join(abstract_parts)
    journal = normalize_space(article.findtext(".//Journal/Title", default=""))

    pub_types = [normalize_space(node.text or "") for node in article.findall(".//PublicationType")]
    article_ids = {}
    for node in article.findall("./PubmedData/ArticleIdList/ArticleId"):
        id_type = node.attrib.get("IdType")
        if id_type:
            article_ids[id_type] = normalize_space(node.text or "")
    doi = article_ids.get("doi", "")
    pmc = article_ids.get("pmc", "")
    databanks = parse_databanks(article)

    pub_date = extract_pub_date(article)
    return {
        "pmid": pmid,
        "title": title,
        "abstract": abstract,
        "journal": journal,
        "publication_types": pub_types,
        "doi": doi,
        "pmc": pmc,
        "databanks": databanks,
        "publish_date": pub_date,
    }


def parse_databanks(article: ET.Element) -> List[Dict[str, object]]:
    databanks = []
    for node in article.findall(".//DataBank"):
        bank_name = normalize_space(node.findtext("DataBankName", default=""))
        accessions = [
            normalize_space(acc.text or "")
            for acc in node.findall(".//AccessionNumber")
            if normalize_space(acc.text or "")
        ]
        if bank_name or accessions:
            databanks.append({"bank_name": bank_name, "accessions": accessions})
    return databanks


def extract_pub_date(article: ET.Element) -> str:
    candidates = article.findall(".//PubDate")
    for node in candidates:
        year = node.findtext("Year")
        month = node.findtext("Month")
        day = node.findtext("Day")
        medline = node.findtext("MedlineDate")
        if year and month:
            month_num = parse_month(month)
            day_num = int(day) if day and day.isdigit() else 1
            return f"{int(year):04d}-{month_num:02d}-{day_num:02d}"
        if medline:
            match = re.search(r"(\d{4})\s+([A-Za-z]{3,9})\s*(\d{1,2})?", medline)
            if match:
                year_num = int(match.group(1))
                month_num = parse_month(match.group(2))
                day_num = int(match.group(3) or "1")
                return f"{year_num:04d}-{month_num:02d}-{day_num:02d}"
    return ""


def parse_month(value: str) -> int:
    lookup = {
        "jan": 1,
        "feb": 2,
        "mar": 3,
        "apr": 4,
        "may": 5,
        "jun": 6,
        "jul": 7,
        "aug": 8,
        "sep": 9,
        "oct": 10,
        "nov": 11,
        "dec": 12,
    }
    key = value.strip().lower()[:3]
    return lookup.get(key, 1)


def is_review(title: str, publication_types: List[str]) -> bool:
    joined = f"{title} {' '.join(publication_types)}".lower()
    return any(re.search(pattern, joined) for pattern in REVIEW_PATTERNS)


def is_cancer_related(text: str) -> bool:
    title, abstract = split_title_abstract(text)
    title_lower = title.lower()
    abstract_lower = abstract.lower()
    title_signal = count_pattern_hits(title_lower, STRONG_CANCER_PATTERNS)
    abstract_signal = count_pattern_hits(abstract_lower, STRONG_CANCER_PATTERNS)
    if title_signal >= 1:
        return True
    return abstract_signal >= 2


def split_title_abstract(text: str) -> Tuple[str, str]:
    if ". " in text:
        return text.split(". ", 1)
    return text, text


def count_pattern_hits(text: str, patterns: List[str]) -> int:
    return sum(len(re.findall(pattern, text)) for pattern in patterns)


def detect_cancer_type(text: str) -> str:
    lowered = text.lower()
    hits = [name for name, patterns in CANCER_PATTERNS.items() if any(re.search(pattern, lowered) for pattern in patterns)]
    return "；".join(hits[:3]) if hits else "未明确"


def classify_omics(text: str) -> Tuple[List[str], str]:
    lowered = text.lower()
    matched = [label for label, patterns in OMICS_PATTERNS.items() if any(re.search(pattern, lowered) for pattern in patterns)]
    explicit = any(re.search(pattern, lowered) for pattern in EXPLICIT_MULTIOMICS_PATTERNS)
    if explicit and "transcriptomics" not in matched and ("single-cell" in lowered or "spatial" in lowered):
        matched.append("transcriptomics")
    matched = list(dict.fromkeys(matched))
    reason = "explicit_multiomics" if explicit else ("multi_layer_detected" if len(matched) >= 2 else "insufficient_layers")
    return matched, reason


def qualifies_multiomics(title: str, abstract: str) -> Tuple[bool, List[str], str]:
    text = f"{title}. {abstract}"
    matched, reason = classify_omics(text)
    lowered = text.lower()
    if any(re.search(pattern, lowered) for pattern in EXPLICIT_MULTIOMICS_PATTERNS):
        if not matched:
            matched = ["multiomics"]
        return True, matched, reason
    if "spatial multi-omics" in lowered or "single-cell multiomics" in lowered:
        if not matched:
            matched = ["multiomics"]
        return True, matched, "explicit_multiomics"
    if len(matched) >= 2:
        return True, matched, reason
    return False, matched, reason


def resolve_doi(session: requests.Session, doi: str) -> Tuple[str, str]:
    if not doi:
        return "", ""
    doi_url = f"https://doi.org/{doi}"
    try:
        response = session.get(doi_url, timeout=30, allow_redirects=True)
        final_url = response.url
        return final_url, doi_url
    except requests.RequestException:
        return doi_url, doi_url


def discover_page_links(session: requests.Session, url: str) -> Tuple[str, str]:
    if not url:
        return "", ""
    try:
        response = session.get(url, timeout=30)
        response.raise_for_status()
    except requests.RequestException:
        return "", ""

    soup = BeautifulSoup(response.text, "html.parser")
    pdf_url = ""
    supplement_url = ""
    for tag in soup.find_all(["a", "meta"]):
        href = tag.get("href") or tag.get("content") or ""
        if not href:
            continue
        href = href.strip()
        joined = requests.compat.urljoin(url, href)
        lowered = joined.lower()
        if not pdf_url and (".pdf" in lowered or "pdf" in lowered):
            pdf_url = joined
        if not supplement_url and any(token in lowered for token in ["supplement", "supplementary", "suppl"]):
            supplement_url = joined
        if pdf_url and supplement_url:
            break
    return pdf_url, supplement_url


def summarize_article(record: Dict[str, object], article_type: str, cancer_type: str, omics_layers: List[str]) -> str:
    text = normalize_space(str(record["abstract"]))
    title = normalize_space(str(record["title"]))
    cancer_cn = cancer_type if cancer_type != "未明确" else "相关肿瘤"
    layers_cn = "、".join(translate_omics_layers(omics_layers)) if omics_layers else "多组学"
    focus = build_focus_phrase(title)
    findings = infer_findings_phrase(text, article_type, cancer_cn)
    data_hint = infer_data_hint(text, layers_cn)
    return f"这篇{article_type}聚焦{cancer_cn}，核心关注点是{focus}。{findings} {data_hint}".strip()


def translate_omics_layers(layers: List[str]) -> List[str]:
    mapping = {
        "genomics": "基因组",
        "epigenomics": "表观组",
        "transcriptomics": "转录组",
        "proteomics": "蛋白组",
        "metabolomics": "代谢/脂质组",
        "microbiome": "微生物组",
        "multiomics": "多组学（原文未细分）",
    }
    return [mapping.get(layer, layer) for layer in layers]


def same_journal(expected: str, actual: str) -> bool:
    def norm(value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", value.lower())

    return norm(expected) == norm(actual)


def build_focus_phrase(title: str) -> str:
    cleaned = re.sub(r"[:;,.]+$", "", title)
    primary = re.split(r"\b(?:using|through|via|with|by|to)\b", cleaned, maxsplit=1, flags=re.IGNORECASE)[0]
    primary = normalize_space(primary)
    return primary[:80] if primary else "该研究主题"


def infer_findings_phrase(text: str, article_type: str, cancer_cn: str) -> str:
    lowered = text.lower()
    if article_type == "综述":
        return f"内容上更偏向梳理{cancer_cn}领域的关键进展、技术整合方式和可转化方向。"
    if "predict" in lowered or "outcome" in lowered or "response" in lowered:
        return f"摘要提示作者重点在做分层预测、疗效判断或结局相关分析，而不只是单纯描述现象。"
    if "immune" in lowered or "immunotherapy" in lowered or "t cell" in lowered:
        return f"摘要信息显示文章较强调免疫微环境、免疫应答或治疗相关机制。"
    if "target" in lowered or "therapy" in lowered or "therapeutic" in lowered:
        return f"摘要信息显示作者不仅做了多组学整合，还尝试把结果落到潜在靶点或治疗启发上。"
    return f"整体上它更像是一篇利用整合数据去解释肿瘤异质性、机制或分型的研究。"


def infer_data_hint(text: str, layers_cn: str) -> str:
    lowered = text.lower()
    if "spatial" in lowered and "single-cell" in lowered:
        return f"从摘要看，核心数据框架包含空间与单细胞层面的联合信息，并由{layers_cn}共同支撑。"
    if "methyl" in lowered and ("rna" in lowered or "transcript" in lowered):
        return f"从摘要看，关键整合点在表观层与表达层的联动分析，核心维度包括{layers_cn}。"
    return f"从摘要可见，这篇文章的主轴是把{layers_cn}放到同一个分析框架里做综合解释。"


def extract_sample_patient_info(text: str) -> str:
    snippets: List[str] = []
    lowered = normalize_space(text)
    for pattern in SAMPLE_PATTERNS:
        for match in re.finditer(pattern, lowered, flags=re.IGNORECASE):
            snippet = lowered[max(0, match.start() - 30): min(len(lowered), match.end() + 40)]
            snippet = normalize_space(snippet).strip(" ,;:.")
            if snippet and snippet not in snippets:
                snippets.append(snippet)
    if not snippets:
        return "摘要未明确说明"
    translated = []
    for snippet in snippets[:4]:
        snippet_cn = snippet
        snippet_cn = re.sub(r"\bpatients?\b", "例患者", snippet_cn, flags=re.IGNORECASE)
        snippet_cn = re.sub(r"\bcases?\b", "例病例", snippet_cn, flags=re.IGNORECASE)
        snippet_cn = re.sub(r"\bsamples?\b", "份样本", snippet_cn, flags=re.IGNORECASE)
        snippet_cn = re.sub(r"\blesions?\b", "个病灶", snippet_cn, flags=re.IGNORECASE)
        snippet_cn = re.sub(r"\bcohorts?\b", "个队列", snippet_cn, flags=re.IGNORECASE)
        snippet_cn = re.sub(r"\bparticipants?\b", "名受试者", snippet_cn, flags=re.IGNORECASE)
        snippet_cn = re.sub(r"\bcells?\b", "个细胞", snippet_cn, flags=re.IGNORECASE)
        translated.append(snippet_cn)
    return "；".join(translated)


def collect_accessions(record: Dict[str, object]) -> List[Tuple[str, str]]:
    found: List[Tuple[str, str]] = []
    seen: Set[Tuple[str, str]] = set()
    for databank in record.get("databanks", []):
        bank_name = str(databank.get("bank_name", "")).strip()
        for accession in databank.get("accessions", []):
            key = (bank_name, accession)
            if accession and key not in seen:
                seen.add(key)
                found.append(key)
    text = f"{record.get('title', '')} {record.get('abstract', '')}"
    for pattern in ACCESSION_PATTERNS:
        for match in re.findall(pattern, text, flags=re.IGNORECASE):
            accession = match.upper()
            key = ("text-mined", accession)
            if key not in seen:
                seen.add(key)
                found.append(key)
    return found


def accession_url(accession: str) -> str:
    upper = accession.upper()
    if upper.startswith(("GSE", "GSM", "GPL", "GDS")):
        return f"https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc={upper}"
    if upper.startswith(("SRP", "SRX", "SRR", "SRS", "ERP", "ERX", "ERR")):
        return f"https://www.ncbi.nlm.nih.gov/sra?term={upper}"
    if upper.startswith(("PRJNA", "PRJEB")):
        return f"https://www.ncbi.nlm.nih.gov/bioproject/?term={upper}"
    if upper.startswith(("EGAS", "EGAD")):
        return f"https://ega-archive.org/search?query={upper}"
    if upper.startswith("PHS"):
        return f"https://www.ncbi.nlm.nih.gov/gap/?term={upper}"
    if upper.startswith("PXD"):
        return f"https://www.ebi.ac.uk/pride/archive/projects/{upper}"
    if upper.startswith("E-MTAB-"):
        return f"https://www.ebi.ac.uk/biostudies/arrayexpress/studies/{upper}"
    return ""


def geo_note(accession: str) -> str:
    upper = accession.upper()
    if upper.startswith("GSE"):
        return f"{upper}=GEO Series"
    if upper.startswith("GSM"):
        return f"{upper}=GEO Sample"
    if upper.startswith("GPL"):
        return f"{upper}=GEO Platform"
    if upper.startswith("GDS"):
        return f"{upper}=GEO Dataset"
    return ""


def build_data_fields(record: Dict[str, object]) -> Dict[str, str]:
    accession_pairs = collect_accessions(record)
    accession_texts = []
    repositories = []
    geo_accessions = []
    geo_notes = []
    download_urls = []

    for source_name, accession in accession_pairs:
        label = f"{accession} ({source_name})" if source_name and source_name != "text-mined" else accession
        accession_texts.append(label)
        if source_name and source_name != "text-mined" and source_name not in repositories:
            repositories.append(source_name)
        url = accession_url(accession)
        if url and url not in download_urls:
            download_urls.append(url)
        note = geo_note(accession)
        if note:
            geo_accessions.append(accession)
            geo_notes.append(note)

    data_downloadable = "是" if accession_pairs else "摘要/元数据未明确"
    return {
        "data_downloadable": data_downloadable,
        "data_repositories": "；".join(repositories) if repositories else "未明确",
        "data_accessions": "；".join(accession_texts) if accession_texts else "",
        "geo_accessions": "；".join(geo_accessions),
        "geo_accession_notes": "；".join(geo_notes),
        "data_download_url": " ; ".join(download_urls),
    }


def safe_sheet_name(name: str) -> str:
    name = re.sub(r"[:\\\\/?*\\[\\]]", "_", name)
    return name[:31]


def set_hyperlink(cell, value: str) -> None:
    if value:
        cell.value = value
        cell.hyperlink = value
        cell.font = Font(color="0000EE", underline="single")
    else:
        cell.value = ""


def write_workbook(rows: List[Dict[str, str]], source_rows: List[Dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    headers = [
        "report_month",
        "journal",
        "publish_date",
        "title",
        "article_type",
        "cancer_type",
        "omics_layers",
        "article_url",
        "doi",
        "abstract_summary_cn",
        "sample_patient_info_cn",
        "data_downloadable",
        "data_repositories",
        "data_accessions",
        "geo_accessions",
        "geo_accession_notes",
        "data_download_url",
        "pubmed_url",
    ]
    url_headers = {
        "article_url",
        "data_download_url",
        "pubmed_url",
    }

    all_sheet = workbook.create_sheet("All_Articles")
    write_sheet(all_sheet, headers, rows, url_headers)

    grouped: Dict[str, List[Dict[str, str]]] = {}
    for row in rows:
        grouped.setdefault(row["journal"], []).append(row)
    for journal, journal_rows in grouped.items():
        sheet = workbook.create_sheet(safe_sheet_name(journal))
        write_sheet(sheet, headers, journal_rows, url_headers)

    source_headers = [
        "journal",
        "publisher",
        "official_toc_url",
        "latest_url",
        "search_url",
        "pubmed_journal_name",
        "query_status",
        "candidate_count",
        "included_count",
        "last_checked_at",
    ]
    source_sheet = workbook.create_sheet("Sources")
    write_sheet(source_sheet, source_headers, source_rows, {"official_toc_url", "latest_url", "search_url"})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_sheet(sheet, headers: List[str], rows: List[Dict[str, str]], url_headers: set) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row_index in range(2, sheet.max_row + 1):
        for col_index, header in enumerate(headers, start=1):
            if header in url_headers:
                value = str(sheet.cell(row=row_index, column=col_index).value or "")
                if value and " ; " not in value and ";" not in value:
                    set_hyperlink(sheet.cell(row=row_index, column=col_index), value)
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 80)
    sheet.freeze_panes = "A2"


def record_in_window(publish_date: str, start_date: dt.date, end_date: dt.date) -> bool:
    if not publish_date:
        return False
    try:
        date_value = dt.date.fromisoformat(publish_date)
    except ValueError:
        return False
    return start_date <= date_value <= end_date


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate a cancer multi-omics monthly journal report.")
    parser.add_argument("--days", type=int, default=30)
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--report-month", default="")
    parser.add_argument("--journal-limit", type=int, default=0)
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--today", default="")
    args = parser.parse_args()

    run_date = dt.date.fromisoformat(args.today) if args.today else today_utc()
    start_date = run_date - dt.timedelta(days=args.days)
    report_month = args.report_month or run_date.strftime("%Y-%m")
    output_path = Path(args.output_dir) / f"journal_cancer_multiomics_{report_month}.xlsx"

    if output_path.exists() and not args.force:
        print(f"SKIP {output_path}")
        return 0

    registry = load_registry()
    if args.journal_limit:
        registry = registry[: args.journal_limit]

    session = build_session()
    all_rows: List[Dict[str, str]] = []
    source_rows: List[Dict[str, str]] = []
    seen_keys = set()

    for journal in registry:
        candidate_count = 0
        included_count = 0
        status = "ok"
        try:
            pmids = pubmed_search(session, journal["pubmed_journal_name"], start_date, run_date)
            candidate_count = len(pmids)
            records = pubmed_fetch_records(session, pmids) if pmids else []
            for record in records:
                combined_text = f"{record['title']} {record['abstract']}"
                if not same_journal(journal["pubmed_journal_name"], str(record["journal"])):
                    continue
                if not record_in_window(str(record["publish_date"]), start_date, run_date):
                    continue
                if not is_cancer_related(combined_text):
                    continue
                qualified, omics_layers, _ = qualifies_multiomics(str(record["title"]), str(record["abstract"]))
                if not qualified:
                    continue

                article_type = "综述" if is_review(str(record["title"]), list(record["publication_types"])) else "原创研究"
                cancer_type = detect_cancer_type(combined_text)
                official_url, doi_url = resolve_doi(session, str(record["doi"]))
                if not official_url:
                    official_url = doi_url
                pubmed_url = f"https://pubmed.ncbi.nlm.nih.gov/{record['pmid']}/" if record["pmid"] else ""

                key = str(record["doi"] or record["pmid"] or record["title"]).lower()
                if key in seen_keys:
                    continue
                seen_keys.add(key)

                data_fields = build_data_fields(record)
                row = {
                    "report_month": report_month,
                    "journal": journal["journal_name"],
                    "publish_date": str(record["publish_date"]),
                    "title": str(record["title"]),
                    "article_type": article_type,
                    "cancer_type": cancer_type,
                    "omics_layers": "、".join(translate_omics_layers(omics_layers)),
                    "article_url": official_url or doi_url or pubmed_url,
                    "doi": str(record["doi"]),
                    "abstract_summary_cn": summarize_article(record, article_type, cancer_type, omics_layers),
                    "sample_patient_info_cn": extract_sample_patient_info(str(record["abstract"])),
                    "pubmed_url": pubmed_url,
                }
                row.update(data_fields)
                all_rows.append(row)
                included_count += 1
        except Exception as exc:  # noqa: BLE001
            status = f"error: {type(exc).__name__}"

        source_rows.append(
            {
                "journal": journal["journal_name"],
                "publisher": journal["publisher"],
                "official_toc_url": journal["official_toc_url"],
                "latest_url": journal["latest_url"],
                "search_url": journal["search_url"],
                "pubmed_journal_name": journal["pubmed_journal_name"],
                "query_status": status,
                "candidate_count": str(candidate_count),
                "included_count": str(included_count),
                "last_checked_at": dt.datetime.now().isoformat(timespec="seconds"),
            }
        )

    all_rows.sort(key=lambda row: (row["publish_date"], row["journal"], row["title"]), reverse=True)
    write_workbook(all_rows, source_rows, output_path)
    print(f"WROTE {output_path}")
    print(f"ARTICLES {len(all_rows)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
